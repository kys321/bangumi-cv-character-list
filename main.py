
#使用样例:python .\main.py --id 34198 7575

import json
import time

from pathlib import Path
import logging
import requests
from tqdm import tqdm
import pandas as pd
import requests
import os
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from auth import do_auth
from utils import env_in_github_workflow
from argparse import ArgumentParser
from PIL import Image as PILImage
from io import BytesIO
API_SERVER = "https://api.bgm.tv"
LOAD_WAIT_MS = 500
ACCESS_TOKEN = ""
IN_GITHUB_WORKFLOW = env_in_github_workflow()

def ensure_user_directory(name):
    """
    确保存在以 name 命名的目录。
    如果目录不存在，则创建它。
    """


    user_dir = os.path.join(name)
    if not os.path.exists(user_dir):
        os.makedirs(user_dir)
        print(f"已创建目录: {user_dir}")

def find_release_date(info_list, key_name):
    # 解决get_update_time查找不到date的问题
    for item in info_list:
        
        if item['key'] == key_name:

            if isinstance(item['value'], list):
                # 如果'value'是列表，提取列表中第一个字典的'v'键值
                return item['value'][0]['v'].split('(')[0]
            elif isinstance(item['value'], str):
                # 如果'value'是字符串，直接返回字符串值
                return item['value']
            break  
    return "未知日期" 

def get_json_with_bearer_token(url):
    time.sleep(LOAD_WAIT_MS/1000)
    logging.debug(f"load url: {url}")
    headers = {'Authorization': 'Bearer ' + ACCESS_TOKEN, 'accept': 'application/json', 'User-Agent': 'bangumi-cv-list/v1'}
    response = requests.get(url, headers=headers)
    
    return response.json()



def load_user_collections(user_id, name):
    user_dir = name  # 使用传入的name作为目录名称
    endpoint = f"{API_SERVER}/v0/persons/{user_id}/characters"
    collections = get_json_with_bearer_token(endpoint)
    
    # 构建新的文件路径
    file_path = os.path.join(user_dir, f"{user_id}_collections.json")
    with open(file_path, "w", encoding="u8") as f:
        json.dump(collections, f, ensure_ascii=False, indent=4)
    
    return collections
def load_user():
    # global USERNAME_OR_UID

    logging.info("loading user info")
    endpoint = f"{API_SERVER}/v0/persons/{id}"

    user_data = get_json_with_bearer_token(endpoint)
    # USERNAME_OR_UID = user_data["username"]
    # print("user_data['username'] =", user_data['username'])
    
    return user_data

    
def trigger_auth():
    global ACCESS_TOKEN
    if IN_GITHUB_WORKFLOW:

        ACCESS_TOKEN = os.environ['BANGUMI_ACCESS_TOKEN']
        return  
    do_auth()
    
    with open("./.bgm_token", "r", encoding="u8") as f:
        tokens = json.load(f)
        ACCESS_TOKEN = tokens["access_token"]
        logging.info("access token loaded")

    if not ACCESS_TOKEN:
        logging.error("ACCESS_TOKEN is empty!")
        raise Exception("need access token (auth failed?)")
    
def characters_info(collections):
    subject_name = []
    subject_id = []
    staff = []
    name = []
    link = []
    for chara in collections:
        subject_name.append(f'{chara["subject_name"]}')
        subject_id.append(f'{chara["subject_id"]}')
        staff.append(f'{chara["staff"]}')
        name.append(f'{chara["name"]}')
        link.append(f'https://bangumi.tv/subject/{chara["subject_id"]}')
    return subject_id, subject_name, staff, name, link




def get_update_time(subject_id_list,user_id,name):
    
    kaifa = []
    update_time = {}
    filename = os.path.join(name, f'{user_id}_kaifa_time.json')

    if os.path.exists(filename):
        print(f'从本地{user_id}_kaifa_time.json读取会社与发行信息')
        with open(filename, 'r',encoding = 'utf-8') as file:
            data = json.load(file)
        
        return data['update_time'], data['kaifa']
    # 使用tqdm包装subject_id_list来创建一个进度条
    for n, subject_id in enumerate(tqdm(subject_id_list, desc="Processing subjects")):
        subject_url = f"{API_SERVER}/v0/subjects/{subject_id}"
        for attempt in range(5):  # 尝试最多5次
            try:
                subject_info = get_json_with_bearer_token(subject_url)
                update_time[n] = subject_info.get("date", "none")
                if update_time[n] is None:
                    update_time[n] = find_release_date(subject_info['infobox'], '发行日期')
                #{'key': '发行日期', 'value': [{'v': '2013-01-18(A-1)'}, {'v': '2013-07-12(A-2)'}, {'v': '2014-02-28(A-3)'}, {'v': '2014-11-21(A-4)'}, {'v': '2016-04-26(A-5)'}]}
                #{'key': '发行日期', 'value': '2019年1月予定'}
                #{'key': '发行日期', 'value': [{'v': '2015-10-31（PSP）'}, {'v': '2024-02-29（Nintendo Switch / Steam）'}]}
                
                
                subject_plat = subject_info.get("platform", "none")
                if subject_plat == "游戏":
                    # {'key': '游戏发行商', 'value': 'Lump of Sugar'}
                    # {'key': '开发', 'value': 'TYPE-MOON'} 'infobox': [{'key': '中文名', 'value': '魔法使之夜'}
                    developer_info = next((item["value"] for item in subject_info.get("infobox", []) if item["key"] == "开发"), None)
                    if developer_info is None:
                        developer_info = next((item["value"] for item in subject_info.get("infobox", []) if item["key"] == "游戏发行商"), "信息缺失")
                    else:
                        developer_info = developer_info
                    if developer_info == '信息缺失':
                        pass
                    kaifa.append(developer_info)
                else:
                    kaifa.append(subject_plat)
                break  # 如果成功，跳出尝试循环
            except Exception as e:
                if attempt == 4:  # 如果是最后一次尝试，记录错误
                    print(f"处理{subject_url}时发生错误：{e}")
                    update_time[n] = "错误"
                    kaifa.append("错误")
                else:
                    time.sleep(1)  # 等待1秒钟后重试

    with open(filename, 'w', encoding='utf-8') as file:
        json.dump({'kaifa': kaifa, 'update_time': update_time}, file, ensure_ascii=False, indent=4)

    print(f'存储会社时间信息至{user_id}_kaifa_time.json')
    return update_time, kaifa


# def get_time_json(update_time_list):
#     with open("updata_time.json","w",encoding="u8") as f:
#         json.dump(update_time_list, f, ensure_ascii=False, indent=4)



def image_download(user_id, chara_number, cv_name):
    user_dir = str(cv_name)
    images_dir = os.path.join(user_dir, f"{user_id}_images")
       # 创建图片存储目录。
    if not os.path.exists(images_dir):
        os.makedirs(images_dir)

    with open(os.path.join(user_dir, f"{user_id}_collections.json"), 'r', encoding='utf-8') as f:
        data = json.load(f)
    # 初始化image（图片链接）列表。
    image = []

    # 初始化进度条。
    pbar = tqdm(total=chara_number, desc="Downloading images")

    for n in range(min(chara_number, len(data))):
        chara = data[n]  # 获取角色数据。
        image_url = chara["images"]["small"]  # 获取小图链接。
        image.append(image_url)  # 将图片链接添加到列表中。

        file_path = os.path.join(images_dir, f"{n}.jpg")

        if os.path.exists(file_path):
            # 如果图片文件已经存在，跳过下载步骤
            pbar.update(1)
            continue

        if image_url:  # 如果图片链接存在。
            for attempt in range(5):  # 尝试最多5次
                try:
                    response = requests.get(image_url, stream=True)  # 请求图片。
                    with open(file_path, "wb") as file:  # 打开文件用于写入。
                        for chunk in response.iter_content(chunk_size=1024):
                            if chunk:
                                file.write(chunk)  # 写入文件。
                    break  # 如果成功，跳出尝试循环
                except Exception as e:
                    if attempt == 4:  # 如果是最后一次尝试，记录错误
                        print(f"处理 {image_url} 时发生错误：{e}")
                    else:
                        time.sleep(1)  # 等待1秒钟后重试

        pbar.update(1)

    pbar.close()
    return image



import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def data_to_xlsx(user_id, name, subject_name, kaifa, staff, time, image, link,cv_name):


    time = {str(k): v for k, v in time.items()}
    time_list = [time[str(i)] for i in range(len(name))]

    df = pd.DataFrame()
    df["肖像"] = []
    df["角色名"] = name
    df["作品名"] = subject_name
    df["开发"] = kaifa
    df["主配角"] = staff
    df["登场时间"] = time_list

    # 构建Excel文件保存的新路径
    excel_file_path = os.path.join(str(cv_name), f"{cv_name}.xlsx")
    df.to_excel(excel_file_path, index=False)


    # 载入Excel文件，并添加图片和超链接
   # 载入Excel文件，并添加图片和超链接
    wb = load_workbook(excel_file_path)
    ws = wb['Sheet1']


    # 设定列宽度
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 15
    
# 假设列宽与像素宽度的换算关系为 1:7
    pixels_per_unit = 7  # 这个值需要根据您的实际pixels_per_unit进行调整
    excel_row_height = 100  # Excel行高的单位
    # 转换Excel行高为像素值，大约是行高*像素/单位（该换算系数可能需要根据实际情况调整）
    fixed_row_height_px = excel_row_height
    for n in range(len(name)):
        N = n + 2  # Excel中行号从1开始，标题行为第一行，因此数据行从2开始

        ws.row_dimensions[N].height = excel_row_height
        user_dir = os.path.join(str(cv_name))
        images_dir = os.path.join(user_dir, f"{user_id}_images")
        if image[n]:
            # 打开图片
            with PILImage.open(os.path.join(images_dir, f"{n}.jpg")) as original_image:

                # 获取单元格的宽度，对应于Excel中的列宽
                cell_width_px = int(ws.column_dimensions['A'].width * pixels_per_unit)
                
                # 裁剪或缩放图片
                if original_image.height > fixed_row_height_px:
                    # 如果原始图片高度大于Excel行高，则裁剪
                    original_image = original_image.crop((0, 0, original_image.width, fixed_row_height_px))
                elif original_image.width > cell_width_px:
                    # 如果图片宽度大于单元格宽度，按比例缩小图片
                    aspect_ratio = original_image.height / original_image.width
                    new_height = int(cell_width_px * aspect_ratio)
                    original_image = original_image.resize((cell_width_px, new_height), PILImage.ANTIALIAS)

                # 保存调整后的图片到BytesIO对象
                image_stream = BytesIO()
                original_image.save(image_stream, format='JPEG')
                image_stream.seek(0)

                # 使用openpyxl的Image类加载BytesIO对象中的图片数据
                img = Image(image_stream)

                # 在工作表中插入图片
                ws.add_image(img, f"A{N}")

    for n in range(len(name)):
        N = n + 2
        ws[f"C{N}"].hyperlink = link[n]
    

    # 保存Excel文件
    wb.save(excel_file_path)

def load_cv_name(user_id):

    url = f"{API_SERVER}/v0/persons/{user_id}"
    time.sleep(LOAD_WAIT_MS/1000)
    headers = {'Authorization': 'Bearer ' + ACCESS_TOKEN, 'accept': 'application/json', 'User-Agent': 'bangumi-cv-list/v1'}
    response = requests.get(url, headers=headers)
    
    info = response.json()['infobox']
    for item in info:
        if item['key'] == '简体中文名':
            
            name = item['value']
            break  
        name = user_id#找不到名字了
    
    return name



def main():
    parser = ArgumentParser(description='bangumi cv character list')
    parser.add_argument('--id', type=int, nargs='+', required=True)
    args = parser.parse_args()
    # args.id 现在是一个 ID 列表
    for user_id in args.id:
        
        

        trigger_auth()
        cv_name = load_cv_name(user_id)
        ensure_user_directory(cv_name)

        
        #如果令牌过期（有效期为一周）手动删除.bgm_token
        collections = load_user_collections(user_id, cv_name) 
        
        subject_id, subject_name, staff, name, link = characters_info(collections)
    
        update_time_list, kaifa = get_update_time(subject_id, user_id, cv_name)
        print(f'Image download for user_id {user_id}')
        image = image_download(user_id, len(subject_id),cv_name)
        print(f'Save to {cv_name}.xlsx')
        data_to_xlsx(user_id, name, subject_name, kaifa, staff, update_time_list, image, link, cv_name)
    
    print('All done')

if __name__ == '__main__':
    main()